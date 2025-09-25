#!/usr/bin/env python3
"""
Excel to YAML Converter for DefensePro Configuration
Converts Excel file with multiple tabs into create_vars_test.yml format
"""

from openpyxl import load_workbook
import yaml
import sys
import os
from typing import Dict, List, Any
import argparse
from pathlib import Path

class ExcelToYamlConverter:
    """Convert Excel configuration to YAML format for DefensePro"""
    
    def __init__(self, excel_file: str, output_file: str = "create_vars_test.yml"):
        self.excel_file = excel_file
        self.output_file = output_file
        self.config = {}
        
    def read_worksheet(self, worksheet_name: str) -> Dict[str, Any]:
        """Read a specific worksheet and return headers and data"""
        try:
            wb = load_workbook(self.excel_file, data_only=True)
            if worksheet_name not in wb.sheetnames:
                return {'headers': [], 'data': []}
            
            ws = wb[worksheet_name]
            headers = []
            data = []
            
            # Get headers from first row
            for cell in ws[1]:
                headers.append(cell.value)
            
            # Get data rows (skip header row)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # Skip completely empty rows
                    data.append(list(row))
            
            return {'headers': headers, 'data': data}
        except Exception as e:
            print(f"Error reading worksheet {worksheet_name}: {e}")
            return {'headers': [], 'data': []}
    
    def parse_config_sheet(self):
        """Parse Config sheet for basic settings"""
        sheet_data = self.read_worksheet('Config')
        dp_ips = []
        policy_config = {}
        
        for row in sheet_data['data']:
            if len(row) >= 2:
                setting = str(row[0]).lower() if row[0] else ''
                value = str(row[1]) if row[1] else ''
                
                if setting == 'dp_ip' and value:
                    dp_ips.append(value)
                elif setting.startswith('create_') or setting == 'apply_policies_after_creation':
                    if value.lower() in ['true', 'false']:
                        policy_config[setting] = value.lower() == 'true'
        
        if dp_ips:
            self.config['dp_ip'] = dp_ips
        if policy_config:
            self.config['security_policy_config'] = policy_config
    
    def parse_network_classes_sheet(self):
        """Parse Network_Classes sheet (transposed format)"""
        sheet_data = self.read_worksheet('Network_Classes')
        if not sheet_data['headers'] or len(sheet_data['headers']) < 2:
            return
        
        # Get network class names from columns (skip Parameter and Required columns)
        class_columns = []
        for i, header in enumerate(sheet_data['headers']):
            if header and 'parameter' not in str(header).lower() and 'required' not in str(header).lower():
                class_columns.append((i, header))
        
        if not class_columns:
            return
        
        # Initialize network classes
        netclasses = {}
        for col_idx, class_name in class_columns:
            netclasses[class_name] = {'name': class_name, 'groups': []}
        
        # Collect address/mask pairs
        address_rows = {}
        mask_rows = {}
        
        for row in sheet_data['data']:
            if not row or not row[0]:
                continue
            
            param_name = str(row[0])
            
            if param_name == 'Name':
                # Update actual names from the Name row
                for col_idx, class_name in class_columns:
                    if col_idx < len(row) and row[col_idx]:
                        actual_name = str(row[col_idx])
                        netclasses[class_name]['name'] = actual_name
            
            elif param_name.startswith('Address_'):
                addr_num = param_name.split('_')[1] if '_' in param_name else '1'
                address_rows[addr_num] = row
            
            elif param_name.startswith('Mask_'):
                mask_num = param_name.split('_')[1] if '_' in param_name else '1'
                mask_rows[mask_num] = row
        
        # Combine address/mask pairs
        for addr_num in address_rows:
            if addr_num in mask_rows:
                addr_row = address_rows[addr_num]
                mask_row = mask_rows[addr_num]
                
                for col_idx, class_name in class_columns:
                    if (col_idx < len(addr_row) and col_idx < len(mask_row) and 
                        addr_row[col_idx] and mask_row[col_idx]):
                        
                        address = str(addr_row[col_idx]).strip()
                        mask = str(mask_row[col_idx]).strip()
                        
                        if address and mask:
                            netclasses[class_name]['groups'].append({
                                'address': address,
                                'mask': mask
                            })
        
        # Add only classes with groups
        result = []
        for class_data in netclasses.values():
            if class_data['groups']:
                result.append(class_data)
        
        if result:
            self.config['netclasses'] = result
    
    def parse_transposed_profile_sheet(self, sheet_name: str, config_key: str, param_mapping: Dict[str, str]):
        """Parse transposed profile sheets (BDOS, DNS, HTTPS, OOS)"""
        sheet_data = self.read_worksheet(sheet_name)
        if not sheet_data['headers'] or len(sheet_data['headers']) < 2:
            return
        
        # Get profile names from columns (skip Parameter and Required columns)
        profile_columns = []
        for i, header in enumerate(sheet_data['headers']):
            if header and 'parameter' not in str(header).lower() and 'required' not in str(header).lower():
                profile_columns.append((i, str(header)))
        
        if not profile_columns:
            return
        
        # Initialize profiles
        profiles = {}
        for col_idx, profile_name in profile_columns:
            profiles[profile_name] = {'name': profile_name, 'params': {}}
        
        # Process each parameter row
        for row in sheet_data['data']:
            if not row or not row[0]:
                continue
            
            param_name = str(row[0])
            if param_name == 'Name':
                # Update actual profile names
                for col_idx, profile_name in profile_columns:
                    if col_idx < len(row) and row[col_idx]:
                        actual_name = str(row[col_idx])
                        profiles[profile_name]['name'] = actual_name
                continue
            
            # Map parameter to YAML key
            yaml_key = param_mapping.get(param_name)
            if not yaml_key:
                continue
            
            # Set parameter values for each profile
            for col_idx, profile_name in profile_columns:
                if col_idx < len(row) and row[col_idx] is not None:
                    value = row[col_idx]
                    if isinstance(value, str):
                        value = value.strip()
                        if not value:
                            continue
                    
                    # Convert numeric values
                    if self.should_be_numeric(yaml_key, sheet_name.lower()):
                        try:
                            value = int(value)
                        except (ValueError, TypeError):
                            value = str(value)
                    else:
                        value = str(value)
                    
                    profiles[profile_name]['params'][yaml_key] = value
        
        # Add profiles with parameters to config
        result = []
        for profile_data in profiles.values():
            if profile_data['params']:
                result.append(profile_data)
        
        if result:
            self.config[config_key] = result
    
    def should_be_numeric(self, param_name: str, sheet_type: str) -> bool:
        """Check if parameter should be numeric"""
        numeric_fields = {
            'bdos': ['inbound_traffic', 'outbound_traffic', 'tcp_in_quota', 'udp_in_quota',
                    'icmp_in_quota', 'igmp_in_quota', 'tcp_out_quota', 'udp_out_quota',
                    'icmp_out_quota', 'igmp_out_quota', 'maximum_interval_between_bursts',
                    'learning_suppression_threshold', 'user_defined_rate_limit'],
            'dns': ['expected_qps', 'max_allow_qps', 'learning_suppression_threshold',
                   'a_quota', 'mx_quota', 'ptr_quota', 'aaaa_quota', 'text_quota',
                   'soa_quota', 'naptr_quota', 'srv_quota', 'other_quota',
                   'manual_trigger_act_thresh', 'manual_trigger_term_thresh'],
            'https': ['rate_limit'],
            'oos': ['act_threshold', 'term_threshold', 'idle_state_bandwidth_threshold', 'idle_state_timer']
        }
        return param_name in numeric_fields.get(sheet_type, [])
    
    def parse_connection_limit_sheet(self):
        """Parse Connection_Limit sheet"""
        sheet_data = self.read_worksheet('Connection_Limit')
        cl_protections = []
        cl_profiles = []
        
        for row in sheet_data['data']:
            if len(row) < 2:
                continue
            
            row_type = str(row[0]).lower() if row[0] else ''
            name = str(row[1]) if row[1] else ''
            
            if not name:
                continue
            
            if row_type == 'protection':
                protection = {'name': name}
                
                # Map protection fields (based on Connection_Limit sheet structure)
                if len(row) > 2 and row[2]:  # Protection_Type
                    protection['protection_type'] = str(row[2])
                if len(row) > 3 and row[3]:  # Protocol
                    protection['protocol'] = str(row[3])
                if len(row) > 4 and row[4]:  # Threshold
                    try:
                        protection['threshold'] = str(int(row[4]))
                    except (ValueError, TypeError):
                        protection['threshold'] = str(row[4])
                if len(row) > 5 and row[5]:  # App_Port_Group
                    protection['app_port_group'] = str(row[5])
                if len(row) > 6 and row[6]:  # Tracking_Type
                    protection['tracking_type'] = str(row[6])
                if len(row) > 7 and row[7]:  # Action
                    protection['action'] = str(row[7])
                if len(row) > 8 and row[8]:  # Packet_Report
                    protection['packet_report'] = str(row[8])
                
                if len(protection) > 1:  # Has more than just name
                    cl_protections.append(protection)
            
            elif row_type == 'profile':
                if len(row) > 9 and row[9]:  # Protections column
                    protections_str = str(row[9])
                    protections_list = [p.strip() for p in protections_str.split(',') if p.strip()]
                    
                    if protections_list:
                        cl_profiles.append({
                            'name': name,
                            'protections': protections_list
                        })
        
        if cl_protections:
            self.config['cl_protections'] = cl_protections
        if cl_profiles:
            self.config['cl_profiles'] = cl_profiles
    
    def parse_security_policies_sheet(self):
        """Parse Security_Policies sheet"""
        sheet_data = self.read_worksheet('Security_Policies')
        policies = []
        
        for row in sheet_data['data']:
            if not row or not row[0]:
                continue
            
            policy_name = str(row[0])
            policy = {'policy_name': policy_name}
            
            # Map policy fields based on Security_Policies sheet structure
            field_mapping = [
                (1, 'state'), (2, 'action'), (3, 'priority'), (4, 'src_network'),
                (5, 'dst_network'), (6, 'direction'), (7, 'packet_reporting_status'),
                (8, 'connection_limit_profile'), (9, 'bdos_profile'), 
                (10, 'dns_flood_profile'), (11, 'https_flood_profile'),
                (12, 'signature_protection_profile'), (13, 'ert_attackers_feed_profile'),
                (14, 'geo_feed_profile'), (15, 'out_of_state_profile')
            ]
            
            for idx, field_name in field_mapping:
                if idx < len(row) and row[idx]:
                    value = str(row[idx]).strip()
                    if value:
                        if field_name == 'priority':
                            try:
                                policy[field_name] = str(int(value))
                            except (ValueError, TypeError):
                                policy[field_name] = value
                        else:
                            policy[field_name] = value
            
            if len(policy) > 1:  # Has more than just policy_name
                policies.append(policy)
        
        if policies:
            self.config['security_policies'] = policies
    
    def convert(self):
        """Main conversion method"""
        print(f"Converting {self.excel_file} to {self.output_file}")
        
        # Parse each sheet
        print("Processing Config sheet...")
        self.parse_config_sheet()
        
        print("Processing Network_Classes sheet...")
        self.parse_network_classes_sheet()
        
        print("Processing BDOS sheet...")
        bdos_mapping = {
            'Action': 'action', 'SYN_Flood': 'syn_flood', 'UDP_Flood': 'udp_flood',
            'IGMP_Flood': 'igmp_flood', 'ICMP_Flood': 'icmp_flood',
            'TCP_ACK_FIN_Flood': 'tcp_ack_fin_flood', 'TCP_RST_Flood': 'tcp_rst_flood',
            'TCP_SYN_ACK_Flood': 'tcp_syn_ack_flood', 'TCP_Frag_Flood': 'tcp_frag_flood',
            'UDP_Frag_Flood': 'udp_frag_flood', 'Inbound_Traffic': 'inbound_traffic',
            'Outbound_Traffic': 'outbound_traffic', 'TCP_In_Quota': 'tcp_in_quota',
            'UDP_In_Quota': 'udp_in_quota', 'ICMP_In_Quota': 'icmp_in_quota',
            'IGMP_In_Quota': 'igmp_in_quota', 'TCP_Out_Quota': 'tcp_out_quota',
            'UDP_Out_Quota': 'udp_out_quota', 'ICMP_Out_Quota': 'icmp_out_quota',
            'IGMP_Out_Quota': 'igmp_out_quota', 'Transparent_Optimization': 'transparent_optimization',
            'Packet_Report': 'packet_report', 'Burst_Attack': 'burst_attack',
            'Max_Interval_Between_Bursts': 'maximum_interval_between_bursts',
            'Learning_Suppression_Threshold': 'learning_suppression_threshold',
            'Footprint_Strictness': 'footprint_strictness',
            'UDP_Packet_Rate_Detection_Sensitivity': 'udp_packet_rate_detection_sensitivity',
            'BDOS_Rate_Limit': 'bdos_rate_limit', 'User_Defined_Rate_Limit': 'user_defined_rate_limit',
            'User_Defined_Rate_Limit_Unit': 'user_defined_rate_limit_unit',
            'Adv_UDP_Detection': 'adv_udp_detection'
        }
        self.parse_transposed_profile_sheet('BDOS', 'bdos_profiles', bdos_mapping)
        
        print("Processing DNS sheet...")
        dns_mapping = {
            'Expected_QPS': 'expected_qps', 'Max_Allow_QPS': 'max_allow_qps',
            'Footprint_Strictness': 'footprint_strictness', 'Packet_Report': 'packet_report',
            'Learning_Suppression_Threshold': 'learning_suppression_threshold',
            'A_Quota': 'a_quota', 'MX_Quota': 'mx_quota', 'PTR_Quota': 'ptr_quota',
            'AAAA_Quota': 'aaaa_quota', 'Text_Quota': 'text_quota', 'SOA_Quota': 'soa_quota',
            'NAPTR_Quota': 'naptr_quota', 'SRV_Quota': 'srv_quota', 'Other_Quota': 'other_quota',
            'A_Status': 'a_status', 'MX_Status': 'mx_status', 'PTR_Status': 'ptr_status',
            'AAAA_Status': 'aaaa_status', 'Text_Status': 'text_status', 'SOA_Status': 'soa_status',
            'NAPTR_Status': 'naptr_status', 'SRV_Status': 'srv_status', 'Other_Status': 'other_status',
            'Manual_Trigger': 'manual_trigger', 'Manual_Trigger_Act_Thresh': 'manual_trigger_act_thresh',
            'Manual_Trigger_Term_Thresh': 'manual_trigger_term_thresh'
        }
        self.parse_transposed_profile_sheet('DNS', 'create_dns_profiles', dns_mapping)
        
        print("Processing HTTPS sheet...")
        https_mapping = {
            'Action': 'action', 'Rate_Limit': 'rate_limit',
            'HTTPS_Auth_Suspect_Sources': 'https_authentication_on_suspect_sources',
            'HTTPS_Auth_All_Sources': 'https_authentication_on_all_sources',
            'Rate_Limit_Status': 'rate_limit_status', 'Packet_Report': 'packet_report',
            'Full_Session_Decryption': 'full_session_decryption'
        }
        self.parse_transposed_profile_sheet('HTTPS', 'create_https_profiles', https_mapping)
        
        print("Processing OOS sheet...")
        oos_mapping = {
            'Action': 'action', 'SYN_ACK_Allow': 'syn_ack_allow', 'Packet_Report': 'packet_report',
            'Risk': 'risk', 'Act_Threshold': 'act_threshold', 'Term_Threshold': 'term_threshold',
            'Idle_State': 'idle_state', 'Idle_State_Bandwidth_Threshold': 'idle_state_bandwidth_threshold',
            'Idle_State_Timer': 'idle_state_timer'
        }
        self.parse_transposed_profile_sheet('OOS', 'oos_profiles', oos_mapping)
        
        print("Processing Connection_Limit sheet...")
        self.parse_connection_limit_sheet()
        
        print("Processing Security_Policies sheet...")
        self.parse_security_policies_sheet()
        
        # Write YAML file
        self.write_yaml()
        print(f"Successfully created {self.output_file}")
    
    def write_yaml(self):
        """Write configuration to YAML file"""
        output_path = Path(self.output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(self.output_file, 'w') as f:
            f.write("---\n")
            f.write(f"# Generated from Excel file: {self.excel_file}\n")
            f.write("# Auto-generated YAML configuration for DefensePro setup\n\n")
            
            yaml.dump(self.config, f, default_flow_style=False, sort_keys=False, 
                     allow_unicode=True, indent=2, width=1000)

def main():
    parser = argparse.ArgumentParser(description='Convert Excel file to DefensePro YAML configuration')
    parser.add_argument('excel_file', help='Input Excel file path')
    parser.add_argument('-o', '--output', default='create_vars_test.yml', 
                       help='Output YAML file path (default: create_vars_test.yml)')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.excel_file):
        print(f"Error: Excel file '{args.excel_file}' does not exist")
        sys.exit(1)
    
    converter = ExcelToYamlConverter(args.excel_file, args.output)
    converter.convert()

if __name__ == "__main__":
    main()
