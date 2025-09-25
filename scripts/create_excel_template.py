#!/usr/bin/env python3
"""
Create Excel Template for DefensePro Configuration
Generates a template Excel file with all required tabs and headers
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import argparse
from pathlib import Path

def create_excel_template(output_file: str = "dp_config_template.xlsx"):
    """Create Excel template with all required tabs"""
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define colors for mandatory/optional fields
    mandatory_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red
    optional_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")   # Light blue
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")     # Light gray
    
    # Config tab
    config_ws = wb.create_sheet("Config")
    config_headers = ['Setting', 'Value', 'Description', 'Required']
    config_data = [
        ['dp_ip', '10.105.192.32', 'DefensePro device IP #1', 'MANDATORY'],
        ['dp_ip', '10.105.192.33', 'DefensePro device IP #2', 'OPTIONAL'],
        ['create_network_classes', 'true', 'Create network classes', 'MANDATORY'],
        ['create_cl_profiles', 'true', 'Create connection limit profiles', 'MANDATORY'],
        ['create_out_of_state_profiles', 'true', 'Create OOS profiles', 'MANDATORY'],
        ['create_bdos_profiles', 'true', 'Create BDoS profiles', 'MANDATORY'],
        ['create_dns_profiles', 'true', 'Create DNS profiles', 'MANDATORY'],
        ['create_https_profiles', 'true', 'Create HTTPS profiles', 'MANDATORY'],
        ['create_security_policies', 'true', 'Create security policies', 'MANDATORY'],
        ['apply_policies_after_creation', 'true', 'Apply policies after creation', 'MANDATORY']
    ]
    
    config_ws.append(config_headers)
    for i, header in enumerate(config_headers):
        config_ws.cell(row=1, column=i+1).fill = header_fill
        config_ws.cell(row=1, column=i+1).font = Font(bold=True)
    
    for row_num, row in enumerate(config_data, start=2):
        config_ws.append(row)
        # Color code based on required field
        fill_color = mandatory_fill if row[3] == 'MANDATORY' else optional_fill
        for col in range(1, 4):  # Don't color the Required column
            config_ws.cell(row=row_num, column=col).fill = fill_color
    
    # Add data validation for true/false fields
    bool_validation = DataValidation(type="list", formula1='"true,false"', allow_blank=False)
    config_ws.add_data_validation(bool_validation)
    bool_validation.add("B3:B10")  # Apply to boolean config rows
    
    # Network Classes tab (transposed format)
    network_ws = wb.create_sheet("Network_Classes")
    network_headers = ['Parameter', 'Network_Class_1', 'Network_Class_2', 'Required']
    network_data = [
        ['Name', 'egor_test_net', 'second_net_class', 'MANDATORY'],
        ['Address_1', '10.10.10.0', '192.168.1.0', 'MANDATORY'],
        ['Mask_1', '255.255.255.0', '255.255.255.0', 'MANDATORY'],
        ['Address_2', '10.10.11.0', '192.168.2.0', 'OPTIONAL'],
        ['Mask_2', '255.255.255.0', '255.255.255.0', 'OPTIONAL'],
        ['Address_3', '10.10.12.0', '', 'OPTIONAL'],
        ['Mask_3', '255.255.255.0', '', 'OPTIONAL'],
        ['Address_4', '10.10.13.0', '', 'OPTIONAL'],
        ['Mask_4', '255.255.255.0', '', 'OPTIONAL']
    ]
    
    network_ws.append(network_headers)
    for i, header in enumerate(network_headers):
        network_ws.cell(row=1, column=i+1).fill = header_fill
        network_ws.cell(row=1, column=i+1).font = Font(bold=True)
    
    for row_num, row in enumerate(network_data, start=2):
        network_ws.append(row)
        fill_color = mandatory_fill if row[3] == 'MANDATORY' else optional_fill
        for col in range(1, 4):
            network_ws.cell(row=row_num, column=col).fill = fill_color
    
    # BDOS tab (transposed format)
    bdos_ws = wb.create_sheet("BDOS")
    bdos_headers = ['Parameter', 'BDOS_Profile_50', 'BDOS_Profile_51', 'Required']
    bdos_data = [
        ['Name', 'BDOS_Profile_50', 'BDOS_Profile_51', 'MANDATORY'],
        ['Action', 'report_only', 'report_only', 'MANDATORY'],
        ['SYN_Flood', 'enable', 'enable', 'OPTIONAL'],
        ['UDP_Flood', 'enable', 'enable', 'OPTIONAL'],
        ['IGMP_Flood', 'enable', 'disable', 'OPTIONAL'],
        ['ICMP_Flood', 'enable', 'enable', 'OPTIONAL'],
        ['TCP_ACK_FIN_Flood', 'enable', 'enable', 'OPTIONAL'],
        ['TCP_RST_Flood', 'enable', 'enable', 'OPTIONAL'],
        ['TCP_SYN_ACK_Flood', 'enable', 'enable', 'OPTIONAL'],
        ['TCP_Frag_Flood', 'enable', 'enable', 'OPTIONAL'],
        ['UDP_Frag_Flood', 'enable', 'enable', 'OPTIONAL'],
        ['Inbound_Traffic', '1000000', '1000000', 'OPTIONAL'],
        ['Outbound_Traffic', '500000', '500000', 'OPTIONAL'],
        ['TCP_In_Quota', '80', '80', 'OPTIONAL'],
        ['UDP_In_Quota', '50', '50', 'OPTIONAL'],
        ['ICMP_In_Quota', '10', '10', 'OPTIONAL'],
        ['IGMP_In_Quota', '50', '50', 'OPTIONAL'],
        ['TCP_Out_Quota', '80', '80', 'OPTIONAL'],
        ['UDP_Out_Quota', '50', '50', 'OPTIONAL'],
        ['ICMP_Out_Quota', '10', '10', 'OPTIONAL'],
        ['IGMP_Out_Quota', '50', '50', 'OPTIONAL'],
        ['Transparent_Optimization', 'enable', 'enable', 'OPTIONAL'],
        ['Packet_Report', 'enable', 'enable', 'OPTIONAL'],
        ['Burst_Attack', 'enable', 'enable', 'OPTIONAL'],
        ['Max_Interval_Between_Bursts', '32', '32', 'OPTIONAL'],
        ['Learning_Suppression_Threshold', '10', '10', 'OPTIONAL'],
        ['Footprint_Strictness', 'medium', 'medium', 'OPTIONAL'],
        ['UDP_Packet_Rate_Detection_Sensitivity', 'low', 'low', 'OPTIONAL'],
        ['BDOS_Rate_Limit', 'user_defined', 'user_defined', 'OPTIONAL'],
        ['User_Defined_Rate_Limit', '500', '500', 'OPTIONAL'],
        ['User_Defined_Rate_Limit_Unit', 'mbps', 'mbps', 'OPTIONAL'],
        ['Adv_UDP_Detection', 'enable', 'enable', 'OPTIONAL']
    ]
    
    bdos_ws.append(bdos_headers)
    for i, header in enumerate(bdos_headers):
        bdos_ws.cell(row=1, column=i+1).fill = header_fill
        bdos_ws.cell(row=1, column=i+1).font = Font(bold=True)
    
    for row_num, row in enumerate(bdos_data, start=2):
        bdos_ws.append(row)
        fill_color = mandatory_fill if row[3] == 'MANDATORY' else optional_fill
        for col in range(1, 4):
            bdos_ws.cell(row=row_num, column=col).fill = fill_color
    
    # Add data validation for BDOS fields
    # Action validation
    action_validation = DataValidation(type="list", formula1='"report_only,block_&_report"', allow_blank=False)
    bdos_ws.add_data_validation(action_validation)
    action_validation.add("B3:C3")  # Action row
    
    # Enable/disable validation - split into separate validations
    enable_validation1 = DataValidation(type="list", formula1='"enable,disable"', allow_blank=False)
    bdos_ws.add_data_validation(enable_validation1)
    enable_validation1.add("B4:C13")  # Flood settings
    
    enable_validation2 = DataValidation(type="list", formula1='"enable,disable"', allow_blank=False)
    bdos_ws.add_data_validation(enable_validation2)
    enable_validation2.add("B23:C25")  # Transparent optimization, packet report, burst attack
    
    enable_validation3 = DataValidation(type="list", formula1='"enable,disable"', allow_blank=False)
    bdos_ws.add_data_validation(enable_validation3)
    enable_validation3.add("B33:C33")  # Advanced UDP detection
    
    # Footprint strictness validation
    footprint_validation = DataValidation(type="list", formula1='"low,medium,high"', allow_blank=False)
    bdos_ws.add_data_validation(footprint_validation)
    footprint_validation.add("B28:C28")
    
    # UDP sensitivity validation
    udp_sensitivity_validation = DataValidation(type="list", formula1='"Ignore_or_Disable,low,medium,high"', allow_blank=False)
    bdos_ws.add_data_validation(udp_sensitivity_validation)
    udp_sensitivity_validation.add("B29:C29")
    
    # Rate limit validation
    rate_limit_validation = DataValidation(type="list", formula1='"disable,normal_edge,suspect_edge,user_defined"', allow_blank=False)
    bdos_ws.add_data_validation(rate_limit_validation)
    rate_limit_validation.add("B30:C30")
    
    # Rate limit unit validation
    unit_validation = DataValidation(type="list", formula1='"kbps,mbps,gbps"', allow_blank=False)
    bdos_ws.add_data_validation(unit_validation)
    unit_validation.add("B32:C32")
    
    # DNS tab (transposed format)
    dns_ws = wb.create_sheet("DNS")
    dns_headers = ['Parameter', 'dns_profile_10', 'dns_profile_11', 'Required']
    dns_data = [
        ['Name', 'dns_profile_10', 'dns_profile_11', 'MANDATORY'],
        ['Expected_QPS', '1000', '2000', 'OPTIONAL'],
        ['Max_Allow_QPS', '5000', '10000', 'OPTIONAL'],
        ['Footprint_Strictness', 'high', 'medium', 'OPTIONAL'],
        ['Packet_Report', 'enable', 'disable', 'OPTIONAL'],
        ['Learning_Suppression_Threshold', '25', '30', 'OPTIONAL'],
        ['A_Quota', '50', '45', 'OPTIONAL'],
        ['MX_Quota', '20', '25', 'OPTIONAL'],
        ['PTR_Quota', '12', '15', 'OPTIONAL'],
        ['AAAA_Quota', '15', '12', 'OPTIONAL'],
        ['Text_Quota', '2', '2', 'OPTIONAL'],
        ['SOA_Quota', '1', '1', 'OPTIONAL'],
        ['NAPTR_Quota', '1', '1', 'OPTIONAL'],
        ['SRV_Quota', '1', '1', 'OPTIONAL'],
        ['Other_Quota', '1', '1', 'OPTIONAL'],
        ['A_Status', 'enable', 'enable', 'OPTIONAL'],
        ['MX_Status', 'enable', 'enable', 'OPTIONAL'],
        ['PTR_Status', 'enable', 'enable', 'OPTIONAL'],
        ['AAAA_Status', 'enable', 'enable', 'OPTIONAL'],
        ['Text_Status', 'enable', 'enable', 'OPTIONAL'],
        ['SOA_Status', 'enable', 'enable', 'OPTIONAL'],
        ['NAPTR_Status', 'enable', 'enable', 'OPTIONAL'],
        ['SRV_Status', 'enable', 'enable', 'OPTIONAL'],
        ['Other_Status', 'enable', 'enable', 'OPTIONAL'],
        ['Manual_Trigger', 'disable', 'disable', 'OPTIONAL'],
        ['Manual_Trigger_Act_Thresh', '0', '0', 'OPTIONAL'],
        ['Manual_Trigger_Term_Thresh', '0', '0', 'OPTIONAL']
    ]
    
    dns_ws.append(dns_headers)
    for i, header in enumerate(dns_headers):
        dns_ws.cell(row=1, column=i+1).fill = header_fill
        dns_ws.cell(row=1, column=i+1).font = Font(bold=True)
    
    for row_num, row in enumerate(dns_data, start=2):
        dns_ws.append(row)
        fill_color = mandatory_fill if row[3] == 'MANDATORY' else optional_fill
        for col in range(1, 4):
            dns_ws.cell(row=row_num, column=col).fill = fill_color
    
    # Add DNS data validation - split into separate validations
    dns_enable_validation1 = DataValidation(type="list", formula1='"enable,disable"', allow_blank=False)
    dns_ws.add_data_validation(dns_enable_validation1)
    dns_enable_validation1.add("B5:C5")  # Packet Report field
    
    dns_enable_validation2 = DataValidation(type="list", formula1='"enable,disable"', allow_blank=False)
    dns_ws.add_data_validation(dns_enable_validation2)
    dns_enable_validation2.add("B16:C26")  # Status fields
    
    dns_footprint_validation = DataValidation(type="list", formula1='"low,medium,high"', allow_blank=False)
    dns_ws.add_data_validation(dns_footprint_validation)
    dns_footprint_validation.add("B4:C4")
    
    # HTTPS tab (transposed format)
    https_ws = wb.create_sheet("HTTPS")
    https_headers = ['Parameter', 'https_profile_1', 'https_profile_2', 'Required']
    https_data = [
        ['Name', 'https_profile_1', 'https_profile_2', 'MANDATORY'],
        ['Action', 'report_only', 'block_&_report', 'MANDATORY'],
        ['Rate_Limit', '2000', '3000', 'OPTIONAL'],
        ['HTTPS_Auth_Suspect_Sources', 'enable', 'enable', 'OPTIONAL'],
        ['HTTPS_Auth_All_Sources', 'enable', 'disable', 'OPTIONAL'],
        ['Rate_Limit_Status', 'enable', 'enable', 'OPTIONAL'],
        ['Packet_Report', 'enable', 'disable', 'OPTIONAL'],
        ['Full_Session_Decryption', 'disable', 'disable', 'OPTIONAL']
    ]
    
    https_ws.append(https_headers)
    for i, header in enumerate(https_headers):
        https_ws.cell(row=1, column=i+1).fill = header_fill
        https_ws.cell(row=1, column=i+1).font = Font(bold=True)
    
    for row_num, row in enumerate(https_data, start=2):
        https_ws.append(row)
        fill_color = mandatory_fill if row[3] == 'MANDATORY' else optional_fill
        for col in range(1, 4):
            https_ws.cell(row=row_num, column=col).fill = fill_color
    
    # Add HTTPS data validation
    https_action_validation = DataValidation(type="list", formula1='"report_only,block_&_report"', allow_blank=False)
    https_ws.add_data_validation(https_action_validation)
    https_action_validation.add("B3:C3")
    
    https_enable_validation = DataValidation(type="list", formula1='"enable,disable"', allow_blank=False)
    https_ws.add_data_validation(https_enable_validation)
    https_enable_validation.add("B4:C8")
    
    # OOS tab (transposed format)
    oos_ws = wb.create_sheet("OOS")
    oos_headers = ['Parameter', 'oos_profile_1', 'oos_profile_2', 'Required']
    oos_data = [
        ['Name', 'oos_profile_1', 'oos_profile_2', 'MANDATORY'],
        ['Action', 'report_only', 'block_&_report', 'OPTIONAL'],
        ['SYN_ACK_Allow', 'disable', 'enable', 'OPTIONAL'],
        ['Packet_Report', 'disable', 'enable', 'OPTIONAL'],
        ['Risk', 'high', 'medium', 'OPTIONAL'],
        ['Act_Threshold', '1000', '1500', 'OPTIONAL'],
        ['Term_Threshold', '500', '750', 'OPTIONAL'],
        ['Idle_State', 'disable', 'enable', 'OPTIONAL'],
        ['Idle_State_Bandwidth_Threshold', '1000', '1200', 'OPTIONAL'],
        ['Idle_State_Timer', '30', '45', 'OPTIONAL']
    ]
    
    oos_ws.append(oos_headers)
    for i, header in enumerate(oos_headers):
        oos_ws.cell(row=1, column=i+1).fill = header_fill
        oos_ws.cell(row=1, column=i+1).font = Font(bold=True)
    
    for row_num, row in enumerate(oos_data, start=2):
        oos_ws.append(row)
        fill_color = mandatory_fill if row[3] == 'MANDATORY' else optional_fill
        for col in range(1, 4):
            oos_ws.cell(row=row_num, column=col).fill = fill_color
    
    # Add OOS data validation - split into separate validations
    oos_action_validation = DataValidation(type="list", formula1='"report_only,block_&_report"', allow_blank=False)
    oos_ws.add_data_validation(oos_action_validation)
    oos_action_validation.add("B3:C3")
    
    oos_enable_validation1 = DataValidation(type="list", formula1='"enable,disable"', allow_blank=False)
    oos_ws.add_data_validation(oos_enable_validation1)
    oos_enable_validation1.add("B4:C5")  # SYN_ACK_Allow and Packet_Report
    
    oos_enable_validation2 = DataValidation(type="list", formula1='"enable,disable"', allow_blank=False)
    oos_ws.add_data_validation(oos_enable_validation2)
    oos_enable_validation2.add("B8:C8")  # Idle_State
    
    oos_risk_validation = DataValidation(type="list", formula1='"info,low,medium,high"', allow_blank=False)
    oos_ws.add_data_validation(oos_risk_validation)
    oos_risk_validation.add("B6:C6")
    
    # Connection Limit tab (keep original format as it's different structure)
    cl_ws = wb.create_sheet("Connection_Limit")
    cl_headers = [
        'Type', 'Name', 'Protection_Type', 'Protocol', 'Threshold',
        'App_Port_Group', 'Tracking_Type', 'Action', 'Packet_Report', 
        'Protections', 'Required'
    ]
    cl_data = [
        ['Protection', 'cl_prot_test1', 'cps', 'tcp', 50, 'http', 'dst_ip', 'drop', 'enable', '', 'MANDATORY'],
        ['Protection', 'cl_prot_test2', 'concurrent_connections', 'udp', 100, 'dns', 'src_ip', 'report_only', 'enable', '', 'MANDATORY'],
        ['Profile', 'cl_prof_example', '', '', '', '', '', '', '', 'cl_prot_test1,cl_prot_test2', 'MANDATORY']
    ]
    
    cl_ws.append(cl_headers)
    for i, header in enumerate(cl_headers):
        cl_ws.cell(row=1, column=i+1).fill = header_fill
        cl_ws.cell(row=1, column=i+1).font = Font(bold=True)
    
    for row_num, row in enumerate(cl_data, start=2):
        cl_ws.append(row)
        fill_color = mandatory_fill if row[10] == 'MANDATORY' else optional_fill
        for col in range(1, 11):
            cl_ws.cell(row=row_num, column=col).fill = fill_color
    
    # Security Policies tab (keep original format)
    policy_ws = wb.create_sheet("Security_Policies")
    policy_headers = [
        'Policy_Name', 'State', 'Action', 'Priority', 'Src_Network', 'Dst_Network',
        'Direction', 'Packet_Reporting_Status', 'Connection_Limit_Profile',
        'BDOS_Profile', 'DNS_Flood_Profile', 'HTTPS_Flood_Profile',
        'Signature_Protection_Profile', 'ERT_Attackers_Feed_Profile',
        'Geo_Feed_Profile', 'Out_Of_State_Profile', 'Required'
    ]
    policy_data = [
        ['example_security_policy_1', 'enable', 'report_only', 700, 'any', 'egor_test_net',
         'oneway', 'enable', 'cl_prof_example', 'BDOS_Profile_50', 'dns_profile_10',
         'https_profile_1', 'All-DoS-Shield', '', '', 'oos_profile_1', 'MANDATORY']
    ]
    
    policy_ws.append(policy_headers)
    for i, header in enumerate(policy_headers):
        policy_ws.cell(row=1, column=i+1).fill = header_fill
        policy_ws.cell(row=1, column=i+1).font = Font(bold=True)
        policy_ws.cell(row=1, column=i+1).alignment = Alignment(wrap_text=True)
    
    for row_num, row in enumerate(policy_data, start=2):
        policy_ws.append(row)
        fill_color = mandatory_fill if row[16] == 'MANDATORY' else optional_fill
        for col in range(1, 17):
            policy_ws.cell(row=row_num, column=col).fill = fill_color
    
    # Adjust column widths for all sheets
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 30)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(output_file)
    
    print(f"Excel template created: {output_file}")
    print("\nTemplate features:")
    print("- Transposed format for easy profile comparison")
    print("- Color coding: Light red = MANDATORY, Light blue = OPTIONAL")
    print("- Data validation dropdowns to prevent errors")
    print("- Automatic column width adjustment")
    print("\nTemplate includes the following tabs:")
    print("- Config: Basic configuration and feature flags")
    print("- Network_Classes: Network class definitions")
    print("- BDOS: BDoS profile settings (transposed)")
    print("- DNS: DNS flood profile settings (transposed)")
    print("- HTTPS: HTTPS flood profile settings (transposed)")
    print("- OOS: Out-of-state profile settings (transposed)")
    print("- Connection_Limit: Connection limit protections and profiles")
    print("- Security_Policies: Security policy definitions with profile bindings")

def main():
    parser = argparse.ArgumentParser(description='Create Excel template for DefensePro configuration')
    parser.add_argument('-o', '--output', default='dp_config_template.xlsx',
                       help='Output Excel template file (default: dp_config_template.xlsx)')
    
    args = parser.parse_args()
    create_excel_template(args.output)

if __name__ == "__main__":
    main()
